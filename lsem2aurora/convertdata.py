__doc__ = """
This is a docstring.

"""
__version__ = 0.1

import os
import time
#import operator
import numpy

import CBNetDB
import json
import openpyxl
import pprint


# environment test and setup

assert os.sys.version_info.major == 3, 'Guess what ... sigh'
assert os.environ['CONDA_DEFAULT_ENV'] == 'sandbox10', 'Guess what ... dev shortcuts'


# set current month
CURRENT_MONTH = int(time.strftime('%m'))
CURRENT_YEAR = int(time.strftime('%Y'))

ctime = time.strftime('%Y-%m-%d')
cdir = os.path.dirname(os.path.abspath(os.sys.argv[0]))
workDir = os.path.abspath(os.path.join(cdir, '..', 'noupload', 'lsem2aurora'))
inDir = os.path.abspath(os.path.join(workDir, 'input_data'))
outDir = os.path.abspath(os.path.join(workDir, 'output'))

for p_ in [workDir, inDir, outDir]:
    if not os.path.exists(p_):
        os.makedirs(p_)

# DB test and setup

# Test database
DB_FILE_NAME = 'aimms_equipment.sqlite'
DB_ACTIVE_TABLE = 'equipment'
# Production DB
DB_FILE_NAME = 'aimms_equipment-test.sqlite'
DB_ACTIVE_TABLE = 'equipment'


aucolsm = ["ID", "title","type","resrearch area","research area pt2","science area specified","pilot domain","short description","university","GPS - X","GPS - Y","Keyword 1","Keyword 2","Keyword 3","Keyword 4","Keyword 5","contact name","contact person email"]
aucolso = ["contact person position", "contact person phone", "support staff name", "support staff position", "support staff email", "support staff phone", "Key Researcher name", "Key Researcher position", "Key Researcher e-mail", "Key Researcher phone number", "website", "Street name", "Street number", "Postal Code", "City", "Country", "Link to photo of resource", "Additional Information"]
aucols = aucolsm + aucolso

lscols = ["ID", "Barcode", "Other ID(s)", "Equipment class", "Laser class", "Name of device", "Manufacturer", "Owner Organisation", "Location"]

aimmscols = lscols + aucols[1:]


# Open database
aimmsDB = CBNetDB.DBTools()
if not os.path.exists(os.path.join(workDir, DB_FILE_NAME)):
    aimmsDB.connectSQLiteDB(DB_FILE_NAME, work_dir=workDir)
    aimmsDB.createDBTable('labs', lscols, primary='ID')
    aimmsDB.createDBTable('aurora', aucols, primary='ID')
    aimmsDB.createDBTable('aimms', aimmscols, primary='ID')

    print(lscols)
    print(aucols)
    print(aimmscols)

else:
    aimmsDB.connectSQLiteDB(DB_FILE_NAME, work_dir=workDir)

#aimmsDB.insertData('labs', {'ID':1})
#aimmsDB.insertData('labs', {'ID':2})
#aimmsDB.insertData('labs', {'ID':3})
#aimmsDB.insertData('labs', {'ID':4})

# Get current new index

if len(aimmsDB.getColumns('labs', ['ID'])[0]) == 0 or aimmsDB.getColumns('labs', ['ID'])[0][-1] == 'None':
    IDXVAL = 1
else:
    IDXVAL = int(aimmsDB.getColumns('labs', ['ID'])[0][-1]) + 1

print('IDXVAL', IDXVAL)

# Database ready


# Load data
ls_data_file = "AIMMS_equipment_list_clean.xlsx"

ls_wb = openpyxl.load_workbook(os.path.join(inDir, ls_data_file))
print(ls_wb.sheetnames)
ls_ws = ls_wb[ls_wb.sheetnames[0]]
print(ls_ws.max_column)
print(ls_ws.max_row)

maxcol = 8
maxrow = 21


for r in range(1, maxrow + 1):
#     print('{}{}'.format(openpyxl.utils.get_column_letter(c), r))
#     print(ls_ws['{}{}'.format(c, r)])
#     print(ls_ws.cell(column=c, row=r).value)
    data = {
        'ID': IDXVAL,
        'barcode': ls_ws.cell(column=1, row=r).value,
        'Other ID(s)': ls_ws.cell(column=2, row=r).value,
        'Equipment class': ls_ws.cell(column=3, row=r).value,
        'Laser class': ls_ws.cell(column=4, row=r).value,
        'Name of device': ls_ws.cell(column=5, row=r).value,
        'Manufacturer': ls_ws.cell(column=6, row=r).value,
        'Owner Organisation': ls_ws.cell(column=7, row=r).value,
        'Location': ls_ws.cell(column=8, row=r).value,
    }
    aimmsDB.insertData('labs', data, commit=False)
    aimmsDB.insertData('aimms', data, commit=False)
    aimmsDB.commitDB()
    IDXVAL += 1

lsdb = aimmsDB.getTable("aimms")









#data = {}
#print('SLdata', len(sldata[0]))
#for r_ in range(len(sldata[0])):
    #data[aimmsDB.getRow(DB_ACTIVE_TABLE, 'doi', sldata[0][r_])[0][0]] = {
        #'year': aimmsDB.getRow(DB_ACTIVE_TABLE, 'doi', sldata[0][r_])[0][4].strip(),
        #'title': aimmsDB.getRow(DB_ACTIVE_TABLE, 'doi', sldata[0][r_])[0][6].strip(),
        #'contributors': [a.strip() for a in aimmsDB.getRow(DB_ACTIVE_TABLE, 'doi', sldata[0][r_])[0][7].replace('.,', '.|').split('|')],
        #'corresponding': aimmsDB.getRow(DB_ACTIVE_TABLE, 'doi', sldata[0][r_])[0][8].strip(),
        #'organisations': [a.strip() for a in aimmsDB.getRow(DB_ACTIVE_TABLE, 'doi', sldata[0][r_])[0][9].split(',')],
    #}
#aimmsDB.closeDB()

#with open('data0.json', 'w') as F:
    #json.dump(data, F, indent=1)

